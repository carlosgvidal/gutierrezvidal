from __future__ import annotations
import csv, io, json, re
from pathlib import Path
from openpyxl import load_workbook
from .models import Document

TEXT_KEYS={'text','texto','body','contenido','content','nota','documento','document'}
DATE_KEYS={'date','fecha','published_at','publication_date'}
TITLE_KEYS={'title','titulo','titular','encabezado'}
SOURCE_KEYS={'source','fuente','medio','outlet'}

def _norm_header(x)->str:
    return re.sub(r'\s+','_',str(x or '').strip().casefold())

def _pick(headers:set[str], keys:set[str])->str|None:
    for k in keys:
        if k in headers: return k
    return None

def rows_to_documents(rows:list[dict], prefix='doc')->list[Document]:
    if not rows: return []
    normalized=[]
    for row in rows:
        nr={_norm_header(k):v for k,v in row.items()}
        normalized.append(nr)
    headers=set().union(*(r.keys() for r in normalized))
    text_key=_pick(headers,TEXT_KEYS)
    if text_key is None:
        # Selección estructural: columna textual con mayor longitud mediana aproximada.
        candidates=[]
        for h in headers:
            vals=[str(r.get(h) or '') for r in normalized]
            score=sum(len(v) for v in vals)/max(1,len(vals))
            candidates.append((score,h))
        text_key=max(candidates)[1] if candidates else None
    date_key=_pick(headers,DATE_KEYS); title_key=_pick(headers,TITLE_KEYS); source_key=_pick(headers,SOURCE_KEYS)
    docs=[]
    for i,row in enumerate(normalized,1):
        txt=str(row.get(text_key) or '').strip() if text_key else ''
        if not txt: continue
        metadata={k:v for k,v in row.items() if k not in {text_key,date_key,title_key,source_key}}
        docs.append(Document(document_id=f'{prefix}-{i}',text=txt,title=(str(row.get(title_key)) if title_key and row.get(title_key) is not None else None),source=(str(row.get(source_key)) if source_key and row.get(source_key) is not None else None),date=(str(row.get(date_key)) if date_key and row.get(date_key) is not None else None),metadata=metadata))
    return docs

def ingest_bytes(filename:str,data:bytes)->list[Document]:
    suffix=Path(filename).suffix.lower()
    if suffix in {'.txt','.md'}:
        return [Document(document_id='doc-1',text=data.decode('utf-8-sig',errors='replace'))]
    if suffix=='.csv':
        text=data.decode('utf-8-sig',errors='replace')
        sample=text[:8192]
        try: dialect=csv.Sniffer().sniff(sample,delimiters=',;\t|')
        except csv.Error: dialect=csv.excel
        rows=list(csv.DictReader(io.StringIO(text),dialect=dialect))
        return rows_to_documents(rows)
    if suffix in {'.xlsx','.xlsm'}:
        wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True)
        docs=[]; seq=1
        for ws in wb.worksheets:
            values=ws.iter_rows(values_only=True)
            try: header=next(values)
            except StopIteration: continue
            headers=[str(v or '') for v in header]
            rows=[]
            for vals in values:
                rows.append({headers[i]:vals[i] for i in range(min(len(headers),len(vals)))})
            part=rows_to_documents(rows,prefix=f'doc-{seq}')
            for d in part:
                d.metadata['worksheet']=ws.title
            docs.extend(part); seq+=1
        return docs
    if suffix=='.json':
        obj=json.loads(data.decode('utf-8-sig'))
        if isinstance(obj,list) and all(isinstance(x,dict) for x in obj): return rows_to_documents(obj)
        if isinstance(obj,dict):
            for key in ('documents','records','rows','data'):
                if isinstance(obj.get(key),list) and all(isinstance(x,dict) for x in obj[key]): return rows_to_documents(obj[key])
            text=next((obj[k] for k in obj if _norm_header(k) in TEXT_KEYS and isinstance(obj[k],str)),None)
            if text is not None: return [Document(document_id='doc-1',text=text)]
        raise ValueError('El JSON no contiene una estructura documental reconocible.')
    raise ValueError('Formato de archivo no admitido.')
